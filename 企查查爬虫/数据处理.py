import openpyxl
import csv
import os.path
ext='.xlsx'
data = []
files = [f for f in os.listdir() if os.path.isfile(f) and os.path.splitext(f)[1]==ext]
for file in files:
    dataall = openpyxl.load_workbook(file)
    sheets = dataall.sheetnames
    sheet = dataall[sheets[0]]
    for i in range(4,sheet.max_row + 1):
        data.append([sheet.cell(row=i, column=1).value,sheet.cell(row=i, column=4).value,sheet.cell(row=i, column=18).value,sheet.cell(row=i, column=21).value,sheet.cell(row=i, column=22).value,sheet.cell(row=i, column=23).value,sheet.cell(row=i, column=24).value,sheet.cell(row=i, column=25).value,sheet.cell(row=i, column=26).value,sheet.cell(row=i, column=27).value])
with open('数据.csv','w',newline='') as f:
    f_csv = csv.writer(f)
    for i in data:
        try:
            f_csv.writerow(i)
        except:
            pass