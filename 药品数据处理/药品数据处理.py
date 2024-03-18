import xlrd
import xlwt
from openpyxl import load_workbook
import csv
import os
ext='.xlsx'
ext2='.XLS'
files = [f for f in os.listdir() if os.path.isfile(f) and (os.path.splitext(f)[1]==ext or os.path.splitext(f)[1]==ext2)]
file = files[0]
wb = load_workbook(file)
sheets = wb.worksheets
# data = xlrd.open_workbook(file)
table = sheets[0]
# table = data.sheets()[0]
s1 = []
c1 = []
c2 = []
target1 = int(input('请输入sheet1要对比的是第几列:')) - 1
for i in range(table.max_column):
    # if table.cell_value(0, i) == '批准文号':
    #     target1 = i
    c1.append(table.cell(0, i).value)
for i in range(1,table.max_row):
    r = []
    for j in range(table.ncols):
        a = table.cell(i, j).value
        if j == target1:
            try:
                a = a.replace('国药准字','').replace('批准文号','').replace('津药制字','').replace(' ','')
            except:
                pass
        r.append(a)
    s1.append(r)
print(s1)
s2 = []
table = data.sheets()[1]
target2 = int(input('请输入sheet2要对比的是第几列:')) - 1
for i in range(table.ncols):
    # if table.cell_value(0, i) == '批准文号':
    #     target2 = i
    c2.append(table.cell_value(0, i))
for i in range(1,table.nrows):
    r = []
    for j in range(table.ncols):
        a = table.cell_value(i, j)
        if j == target2:
            try:
                a = a.replace('国药准字','').replace('注册证号','').replace('津药制字','').replace(' ','')
            except:
                pass
        r.append(a)
    s2.append(r)
s3 = []
for i in s1:
    for j in s2:
        if i[target1] == j[target2]:
            s3.append(i)
            break
s4 = []
for i in s2:
    for j in s1:
        if i[target2] == j[target1]:
            s4.append(i)
            break
print(s4)
s3.sort(key=lambda x:x[target1])
s4.sort(key=lambda x:x[target2])
style1 = xlwt.easyxf('')
style2 = xlwt.easyxf('pattern: pattern solid, fore_colour yellow;')
workbook = xlwt.Workbook(encoding = 'utf-8')
sheet1 = workbook.add_sheet('sheet1',cell_overwrite_ok=True)
head = c1[::]
for i in range(len(c1)):
    sheet1.write(0, i, label=c1[i])
for i in range(1,len(s1)+1):
    for j in range(len(c1)):
        sheet1.write(i,j, label = s1[i-1][j])
head = c2[::]
sheet2 = workbook.add_sheet('sheet2',cell_overwrite_ok=True)
for i in range(len(c2)):
    sheet2.write(0, i, label=c2[i])
for i in range(1,len(s2)+1):
    for j in range(len(c2)):
        sheet2.write(i,j, label = s2[i-1][j])
sheet3 = workbook.add_sheet('sheet3',cell_overwrite_ok=True)
for i in range(len(c1)):
    sheet3.write(0, i, label=c1[i])
for i in range(1,len(s3)):
    for j in range(len(c1)):
        try:
            if s3[i-1][target1] == s3[i][target1] or s3[i-1][target1] == s3[i-2][target1]:
                sheet3.write(i,j,label = s3[i-1][j],style = style2)
            else:
                sheet3.write(i, j, label=s3[i-1][j], style=style1)
        except:
            sheet3.write(i, j, label=s3[i-1][j], style=style1)
sheet4 = workbook.add_sheet('sheet4',cell_overwrite_ok=True)
for i in range(len(c2)):
    sheet4.write(0, i, label=c2[i])
for i in range(1,len(s4)):
    for j in range(len(c2)):
        try:
            if s4[i-1][target2] == s4[i][target2] or s4[i-1][target2] == s4[i-2][target2]:
                sheet4.write(i,j,label = s4[i-1][j],style = style2)
            else:
                sheet4.write(i, j, label=s4[i-1][j], style=style1)
        except:
            sheet4.write(i, j, label=s4[i-1][j], style=style1)
workbook.save('结果.xls')