import openpyxl
from collections import Counter
d = ['科技技术公司', '航天技术公司', '学校单位', '汽车公司', '其他单位或企业', '电器电子数码制造', '通讯公司', '金融', '电力公司', '技术公司', '机关单位', '管理服务公司', '互联网公司', '科研单位', '教育机构', '咨询服务公司', '电信运营商', '邮政单位', '银行']
dataall = openpyxl.load_workbook('工作簿1.xlsx')
sheets = dataall.sheetnames
sheet = dataall[sheets[0]]
beiyou = []
bbb = []
for i in range(1,sheet.max_row + 1):
    beiyou.append([sheet.cell(row=i, column=2).value,sheet.cell(row=i, column=4).value,sheet.cell(row=i, column=5).value])
    bbb.append(sheet.cell(row=i, column=2).value)
sheet = dataall[sheets[1]]
chengdian = []
cdb = []
for i in range(1,sheet.max_row + 1):
    chengdian.append([sheet.cell(row=i, column=2).value,sheet.cell(row=i, column=4).value])
    cdb.append(sheet.cell(row=i, column=2).value)
sheet = dataall[sheets[2]]
xidian = []
xdb = []
for i in range(1,sheet.max_row + 1):
    xidian.append([sheet.cell(row=i, column=2).value,sheet.cell(row=i, column=4).value])
    xdb.append(sheet.cell(row=i, column=2).value)
fenlei = []
sheet = dataall[sheets[3]]
for i in range(1,sheet.max_row + 1):
    fenlei.append([sheet.cell(row=i, column=2).value,sheet.cell(row=i, column=3).value])
for i in fenlei:
    if i[0] in bbb:
        i.append('北邮')
        i.append(beiyou[bbb.index(i[0])][1])
    elif i[0] in cdb:
        i.append('成电')
        i.append(chengdian[cdb.index(i[0])][1])
    else:
        i.append('西电')
        i.append(xidian[xdb.index(i[0])][1])
fenlei = fenlei[1:]
beiyou = beiyou[1:]
chengdian = chengdian[1:]
xidian = xidian[1:]
bydic = {}
cddic = {}
xddic = {}
for i in d:
    bydic[i] = 0
    cddic[i] = 0
    xddic[i] = 0
for i in fenlei:
    if i[2] == '北邮':
        if i[1] != '其他':
            bydic[i[1]] += int(i[-1])
    elif i[2] == '成电':
        if i[1] != '其他':
            cddic[i[1]] += int(i[-1])
    elif i[2] == '西电':
        if i[1] != '其他':
            xddic[i[1]] += int(i[-1])
print('北邮雇主类型TOP10为：' + ',   '.join([i[0] + '(' + str(i[1]) + ')' for i in sorted([[i,j] for i,j in bydic.items()],key=lambda x:x[1],reverse=True)[:10]]))
print('西电雇主类型TOP10为：' + ',   '.join([i[0] + '(' + str(i[1]) + ')' for i in sorted([[i,j] for i,j in xddic.items()],key=lambda x:x[1],reverse=True)[:10]]))
print('成电雇主类型TOP10为：' + ',   '.join([i[0] + '(' + str(i[1]) + ')' for i in sorted([[i,j] for i,j in cddic.items()],key=lambda x:x[1],reverse=True)[:10]]))
beiyou1 = sorted(beiyou,key=lambda x:x[1],reverse=True)
chengdian1 = sorted(chengdian,key=lambda x:x[1],reverse=True)
xidian1 = sorted(xidian,key=lambda x:x[1],reverse=True)
print('北邮最关注的招聘TOP20：' + ',   '.join([i[0] + '(' + str(i[1]) + ')'  for i in beiyou1[:20]]))
print('成电最关注的招聘TOP20：' + ',   '.join([i[0] + '(' + str(i[1]) + ')'  for i in chengdian1[:20]]))
print('西电最关注的招聘TOP20：' + ',   '.join([i[0] + '(' + str(i[1]) + ')'  for i in xidian1[:20]]))
byzp = []
for i in beiyou:
    byzp.append([i[0],i[-1]])
byzp.sort(key=lambda x:x[1],reverse=True)
print('北邮招聘数量TOP10:' + ',   '.join([i[0] + '(' + str(i[1]) + ')'  for i in byzp[:10]]))
ict1 = []
for i in fenlei:
    ict1.append(i[0])
ict2 = []
for i,j in Counter(ict1).items():
    if j > 3:
        ict2.append([i,j])
ict3 = []
for i in ict2:
    for j in fenlei:
        if i[0] == j[0]:
            ict3.append([i[0],j[-1]])
            break
print('ICT TOP10:' + ',   '.join([i[0] + '(' + str(i[1]) + ')'  for i in ict3[:10]]))